# |======================================================|
# | (C) Devon Energy Corporation                         |
# | Paul PC - paul.poputa-clean@dvn.com                  |
# | Distributed under the Apache License (i think        |
# | Script to output the timeline as a colorful xls file |
# |======================================================|

from lxml import etree
from lxml.builder import E
import csv
import socket
import codecs
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.style import Color, Fill
from openpyxl.cell import Cell
from openpyxl.cell import get_column_letter


# reads te timeline from a CSV file. Temporary method until we can merge it with plaso and output the xls from the object in memory.
def read_from_csv(filename):
    timeline=[]
    # lacking a better method of cleaning up the file, let't create a temporary one without the null characters
    res=open(filename+"_nn","w")
    for line in open(filename,"r"):
        res.write(line.replace('\0','?'))       
    res.close
    
    with open(filename+"_nn",'rb') as csvfile:
        headers=[]
        #try:
        #    spamreader = csv.reader(unicode_csv, delimiter=',', quotechar='"')
        #except:
        #    print("spamreader is the error")
        spamreader = csv.reader(csvfile, delimiter=',', quotechar='"')
        linenr=1
        # will have to do something to deal with null characters. it will have to be a modifier for the csv function
        for row in spamreader:
            try:
                linenr+=1
                if (len(headers) == 0):
                    headers = row
                else:
                    elements={}
                    for i, element in enumerate(row):
                        elements[headers[i]]=element
                        timeline.append(elements)
            except:
                print("error when reading line"+linenr)
                linenr+=1
    return timeline

# helper method to append the proper formating rule once a match has been found
def add_formatting(field, condition):
    frm={}
    for frm_c in ["bold","font","color","background","id"]:
        if frm_c in condition.keys():
            frm[frm_c]=condition[frm_c]
    return frm

# checks the row for color depending on the values in each column. Returns the formatting for the cell 
def check_colors(row, conditions):
    color_rules={}
    # i am aware that this is a very inefficient way of doing things - once i get it working, we can look at increasing the efficiency of the preprocessing
    for cnd in conditions:
        # checking to see if the conditions apply to the row
        #print row.keys(), cnd["field"]
        if cnd["field"] in row.keys():
            #print "rule" + cnd["id"]+" might apply to "+row["short"]
            if cnd["operator"] == "contain":
                if row[cnd["field"]].upper().find(cnd["text"].upper())>=0:
                    color_rules[cnd["field"]]=add_formatting(cnd["field"],cnd)
            elif cnd["operator"] == "equal":
                if row[cnd["field"]].upper()==cnd["text"].upper():
                    color_rules[cnd["field"]]=add_formatting(cnd["field"],cnd)
            elif cnd["operator"] == "beginwith":
                if row[cnd["field"]].upper().startswith(cnd["text"].upper()):
                    color_rules[cnd["field"]]=add_formatting(cnd["field"],cnd)
            elif cnd["operator"] == "endwith":
                if row[cnd["field"]].upper().endswith(cnd["text"].upper()):
                    color_rules[cnd["field"]]=add_formatting(cnd["field"],cnd)
            else:
                    print "cannot parse for the operator: "+cnd["operator"]
        elif cnd["field"]=="[ALL]":
            for key in row.keys():
                value=row[key]
                if cnd["operator"] == "contain":
                    if value.upper().find(cnd["text"].upper())>=0:
                        color_rules[key]=add_formatting(key,cnd)
                elif cnd["operator"] == "equal":
                    if value.upper()==cnd["text"].upper():
                        color_rules[key]=add_formatting(key,cnd)
                elif cnd["operator"] == "beginwith":
                    if value.upper().startswith(cnd["text"].upper()):
                        color_rules[key]=add_formatting(key,cnd)
                elif cnd["operator"] == "endwith":
                    if value.upper().endswith(cnd["text"].upper()):
                        color_rules[key]=add_formatting(key,cnd)
                else:
                    print "cannot parse for the operator: "+cnd["operator"]                
            
#    if color_rules:
#        print color_rules
    return color_rules
# loads the conditions for the color template
def load_conditions(condfile):
    conditions=[]
    tree=etree.parse(open(condfile,'r'))
    for cond in tree.iter("formula"):
        cnd={}
        cnd["text"]=cond.text
        for attr, value in cond.items():
            if value != "":
                cnd[attr]=value
        conditions.append(cnd)
    return conditions

def CLASS(*args): # class is a reserved word in Python
    return {"class":' '.join(args)}

# exports the timeline as an html file
def export_html(timeline,file,columns):
    c=load_conditions("formating.xml")
    html= page = (
        E.html(
            E.head(
                E.title("Timeline as HTML output")
            ),         
        )
    )
    bdy=etree.SubElement(page,"body")
    tbl=etree.SubElement(bdy,"table")
    tbl.set("border","1")
    tbl_thead=etree.SubElement(tbl,"thead")
    tbl_h=etree.SubElement(tbl_thead,"tr")
    for tlk in columns:
        tbl_td=etree.SubElement(tbl_h,"td")
        tbl_td.text=tlk
    tbl_tbody=etree.SubElement(tbl,"tbody")  
    for row in timeline:
        tbl_tr=etree.SubElement(tbl_tbody,"tr")
        # adding formatting to the row
        frm=check_colors(row,c)
        for col in columns:
            tbl_td=etree.SubElement(tbl_tr,"td")
            try:
                tbl_td.text=row[col]
                if col in frm.keys():
                    tbl_td.set("style","color: "+frm[col]["color"]+"; background-color: "+frm[col]["background"])
            except:
                print("encountered errors when processing: "+row[col])            
    f=open(file,'w')
    f.write(etree.tostring(page, pretty_print=True))
    f.close()

# sends the message to a syslog server
def send_to_splunk(timeline,ip,portnum):
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM, 0)
    s.connect((ip, portnum))
    for row in timeline:
        line=row["date"]+" "+row["time"] + " "+row["source"]+": "+row["short"]+" -- "+row["desc"]
        s.send(line)
    s.close()

# exports into an excel file
def export_excel(timeline,filename,columns):
    wb = Workbook()
    dest_filename = r'Timeline.xlsx'

    ws = wb.worksheets[0]

    ws.title = "Timeline"

    for col_idx in xrange(1, 5):
        col = get_column_letter(col_idx)
    for row in xrange(1, 5):
        _cell = ws.cell('%s%s'%(col, row))
        _cell.value = '%s%s' % (col, row)
  
    # Cell font style
      # Style information can be found in openpyxl/style.py
        _cell.style.font.color.index = Color.GREEN
        _cell.style.font.name = 'Arial'
        _cell.style.font.size = 8
        _cell.style.font.bold = True
        _cell.style.alignment.wrap_text = True
      
      # Cell background color
        _cell.style.fill.fill_type = Fill.FILL_SOLID
        _cell.style.fill.start_color.index = Color.DARKRED
  
      # You should only modify column dimensions after you have written a cell in 
      #     the column. Perfect world: write column dimensions once per column
      # 
      #ws.column_dimensions["F"].width = 60.0


    #ws = wb.create_sheet()

    #ws.title = 'Pi'

    #ws.cell('F5').value = 3.14

    wb.save(filename = dest_filename)


timeline=read_from_csv('timeline_march_3h_135.csv')
#print timeline
#export_html(timeline,"test.html",["date","time","source","short"])
#send_to_splunk(timeline,"127.0.0.1",522)


#for row in timeline:
#    check_colors(row, c)
export_html(timeline,"timeline_march_3h_135.html",["date","time","source","short"])