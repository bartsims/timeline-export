from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.style import Color, Fill
from openpyxl.cell import Cell
from openpyxl.cell import get_column_letter

wb = Workbook()

dest_filename = r'Timeline.xlsx'

ws = wb.worksheets[0]

ws.title = "Timeline"

for col_idx in xrange(1, 5):
  col = get_column_letter(col_idx)
  for row in xrange(1, 5):
    _cell = ws.cell('%s%s'%(col, row))
    _cell.value = '%s%s' % (col, row)
  
    # Cell font style
  	# Style information can be found in openpyxl/style.py
  	_cell.style.font.color.index = Color.GREEN
  	_cell.style.font.name = 'Arial'
  	_cell.style.font.size = 8
  	_cell.style.font.bold = True
  	_cell.style.alignment.wrap_text = True
  	
  	# Cell background color
  	_cell.style.fill.fill_type = Fill.FILL_SOLID
  	_cell.style.fill.start_color.index = Color.DARKRED
  
  	# You should only modify column dimensions after you have written a cell in 
  	#     the column. Perfect world: write column dimensions once per column
  	# 
  	#ws.column_dimensions["F"].width = 60.0


#ws = wb.create_sheet()

#ws.title = 'Pi'

#ws.cell('F5').value = 3.14

wb.save(filename = dest_filename)
